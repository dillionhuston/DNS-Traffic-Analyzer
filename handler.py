import openpyxl
import pandas as pd
import os
from openpyxl import Workbook 
from openpyxl import load_workbook



dns_traffic_data = "data.xlsx"

class HandleData:
    
    def read_data(dns_traffic_data):
        try:
            wb = openpyxl.load_workbook(dns_traffic_data)
            traffic_obj = wb.active  
            print(traffic_obj) 
            for row in traffic_obj.iter_rows(min_row=1, max_row=6, values_only=True):
                print(row)
        except Exception as e:
            print(f"Error reading {dns_traffic_data}: {e}")
            return 
        

